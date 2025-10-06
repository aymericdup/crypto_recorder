import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import glob
warnings.filterwarnings('ignore')

def load_ticker_files(file_pattern, exchange_name):
    """
    Load multiple parquet files for one exchange (one file per ticker)
    
    Args:
        file_pattern (str): Pattern to match files (e.g., "binance_*.parquet" or list of files)
        exchange_name (str): Name of the exchange
        
    Returns:
        pd.DataFrame: Combined dataframe with all tickers
    """
    print(f"Loading {exchange_name} data...")
    
    # Handle both file patterns and file lists
    if isinstance(file_pattern, str):
        if '*' in file_pattern:
            # Glob pattern
            files = glob.glob(file_pattern)
        else:
            # Single file
            files = [file_pattern]
    else:
        # List of files
        files = file_pattern
    
    if not files:
        raise ValueError(f"No files found matching pattern: {file_pattern}")
    
    combined_dfs = []
    
    for file_path in files:
        print(f"  Loading {file_path}...")
        df = pd.read_parquet(file_path)
        
        # Ensure datetime column
        df['datetime'] = pd.to_datetime(df['datetime'])
        df["datetime"] = df["datetime"].dt.floor("H") # Bucket to the hour (drop mins/seconds/micros)
        df.drop_duplicates(subset=['datetime'], inplace=True)
        df = df.sort_values('datetime')
        df['exchange'] = exchange_name
        
        # Extract ticker from filename if not in data
        if 'ticker' not in df.columns:
            # Try to extract ticker from filename
            filename = os.path.basename(file_path)
            # Remove exchange name and extension, assume format like "binance_BTCUSDT.parquet"
            ticker = filename.replace(f"{exchange_name.lower()}_", "").replace(".parquet", "")
            df['ticker'] = ticker.upper()
        
        combined_dfs.append(df)
    
    # Combine all dataframes
    combined_df = pd.concat(combined_dfs, ignore_index=True)
    combined_df = combined_df.sort_values(['ticker', 'datetime'])
    
    print(f"{exchange_name} data: {len(combined_df)} rows, {len(combined_df['ticker'].unique())} tickers")
    print(f"Tickers: {sorted(combined_df['ticker'].unique())}")
    
    return combined_df

def extract_ticker_from_filename(file_path, exchange_name):
    """
    Extract ticker symbol from filename
    
    Args:
        file_path (str): Path to the file
        exchange_name (str): Exchange name to remove from filename
        
    Returns:
        str: Ticker symbol
    """
    filename = os.path.basename(file_path)
    
    # Common patterns for ticker extraction
    patterns = [
        f"{exchange_name.lower()}_",
        f"{exchange_name.upper()}_",
        exchange_name.lower() + "_",
        exchange_name.upper() + "_"
    ]
    
    ticker = filename
    for pattern in patterns:
        ticker = ticker.replace(pattern, "")
    
    # Remove extension
    ticker = ticker.replace(".parquet", "").replace(".csv", "")
    
    return ticker.upper()

def load_exchange_files_from_directory(directory_path, exchange_name, file_extension="*.parquet"):
    """
    Load all files from a directory for one exchange
    
    Args:
        directory_path (str): Path to directory containing files
        exchange_name (str): Name of the exchange
        file_extension (str): File pattern (default: "*.parquet")
        
    Returns:
        pd.DataFrame: Combined dataframe
    """
    pattern = os.path.join(directory_path, file_extension)
    return load_ticker_files(pattern, exchange_name)

def get_common_tickers(df1, df2):
    """Get common tickers between two exchanges"""
    tickers1 = set(df1['ticker'].unique())
    tickers2 = set(df2['ticker'].unique())
    common_tickers = list(tickers1.intersection(tickers2))
    
    print(f"\nExchange 1 tickers: {len(tickers1)} - {sorted(tickers1)}")
    print(f"Exchange 2 tickers: {len(tickers2)} - {sorted(tickers2)}")
    print(f"Common tickers: {len(common_tickers)} - {sorted(common_tickers)}")
    
    if not common_tickers:
        print("WARNING: No common tickers found between exchanges!")
        print("Please check that ticker naming is consistent between files.")
    
    return common_tickers

def synchronize_pair_data(df1, df2, ticker, exchange1_name, exchange2_name, tolerance_minutes=1):
    """
    Synchronize data between exchanges for a specific ticker
    
    Args:
        df1, df2 (pd.DataFrame): Exchange dataframes
        ticker (str): Ticker symbol
        exchange1_name, exchange2_name (str): Exchange names
        tolerance_minutes (int): Time tolerance for matching
        
    Returns:
        pd.DataFrame: Synchronized data
    """
    # Filter data for specific ticker
    data1 = df1[df1['ticker'] == ticker].copy()
    data2 = df2[df2['ticker'] == ticker].copy()
    
    if len(data1) == 0:
        print(f"WARNING: No data found for {ticker} in {exchange1_name}")
        return pd.DataFrame()
    
    if len(data2) == 0:
        print(f"WARNING: No data found for {ticker} in {exchange2_name}")
        return pd.DataFrame()
    
    print(f"  {ticker}: {len(data1)} points from {exchange1_name}, {len(data2)} points from {exchange2_name}")
    
    # Prepare for merge
    data1_reset = data1[['datetime', 'fr', 'ticker']].reset_index(drop=True)
    data2_reset = data2[['datetime', 'fr', 'ticker']].reset_index(drop=True)
    
    # Add premium if available
    if 'premium' in data1.columns: data1_reset['premium'] = data1['premium'].values
    if 'premium' in data2.columns: data2_reset['premium'] = data2['premium'].values
    
    # merge
    synchronized = pd.merge(
        data1_reset.sort_values('datetime'),
        data2_reset.sort_values('datetime'),
        suffixes=(f'_{exchange1_name}', f'_{exchange2_name}'),
        on="datetime",
        how="inner"
    )

    # Merge with tolerance
    # tolerance = pd.Timedelta(minutes=tolerance_minutes)
    
    # synchronized = pd.merge_asof(
        # data1_reset.sort_values('datetime'),
        # data2_reset.sort_values('datetime'),
        # on='datetime',
        # suffixes=(f'_{exchange1_name}', f'_{exchange2_name}'),
        # tolerance=tolerance,
        # direction='nearest'
    # )
    
    # Remove rows where no match was found
    synchronized = synchronized.dropna(subset=[f'fr_{exchange1_name}', f'fr_{exchange2_name}'])
    
    if len(synchronized) > 0:
        # Compute spreads
        synchronized['spread'] = synchronized[f'fr_{exchange1_name}'] - synchronized[f'fr_{exchange2_name}']
        synchronized['spread_abs'] = abs(synchronized['spread'])
        synchronized['spread_pct'] = synchronized['spread'] * 100
        
        # Add relative spread
        avg_fr = (synchronized[f'fr_{exchange1_name}'] + synchronized[f'fr_{exchange2_name}']) / 2
        synchronized['spread_relative'] = (synchronized['spread_abs'] / abs(avg_fr)) * 100
        
        synchronized['ticker'] = ticker
        synchronized['pair_name'] = f"{ticker}_{exchange1_name}_{exchange2_name}"
        
        print(f"  {ticker}: {len(synchronized)} synchronized points")
    else:
        print(f"  {ticker}: No synchronized points found!")
    
    return synchronized

def identify_arbitrage_opportunities(sync_data, ticker, exchange1_name, exchange2_name, threshold_pct=0.5, exit_delay_periods = 3):
    """
    Identify individual arbitrage opportunities above a threshold
    Considers sign changes as separate opportunities (different trading directions)
    
    Args:
        sync_data (pd.DataFrame): Synchronized data
        ticker (str): Ticker symbol
        exchange1_name, exchange2_name (str): Exchange names
        threshold_pct (float): Percentage threshold for opportunities
        
    Returns:
        pd.DataFrame: DataFrame with individual opportunities
    """
    if len(sync_data) == 0:
        return pd.DataFrame()
    
    # Sort by datetime
    data = sync_data.sort_values('datetime').copy()
    
    # Calculate normalized spread (z-score)
    spread_mean = data['spread_pct'].mean()
    spread_std = data['spread_pct'].std()
    data['norm_spread'] = (data['spread_pct'] - spread_mean) / spread_std if spread_std > 0 else 0
    
    # Find opportunities separately for positive and negative spreads
    positive_opportunities = data['spread_pct'] >= threshold_pct  # Exchange1 > Exchange2
    negative_opportunities = data['spread_pct'] <= -threshold_pct  # Exchange2 > Exchange1
    
    opportunities = []
    
    # Process positive opportunities (Exchange1 higher funding rate)
    opportunities.extend(_extract_opportunities(
        data, positive_opportunities, ticker, exchange1_name, exchange2_name, 
        threshold_pct, 'positive', exit_delay_periods))
    
    # Process negative opportunities (Exchange2 higher funding rate)  
    opportunities.extend(_extract_opportunities(
        data, negative_opportunities, ticker, exchange1_name, exchange2_name,
        threshold_pct, 'negative', exit_delay_periods))
    
    # Sort all opportunities by starting date (latest first)
    if opportunities:
        opportunities_df = pd.DataFrame(opportunities)
        opportunities_df = opportunities_df.sort_values('starting_date', ascending=False)
        return opportunities_df
    
    return pd.DataFrame()

def _extract_opportunities(data, opportunity_mask, ticker, exchange1_name, exchange2_name, threshold_pct, direction, exit_delay_periods):
    """
    Extract consecutive opportunities from a boolean mask
    
    Args:
        data (pd.DataFrame): Synchronized data
        opportunity_mask (pd.Series): Boolean series indicating opportunities
        ticker (str): Ticker symbol
        exchange1_name, exchange2_name (str): Exchange names
        threshold_pct (float): Threshold used
        direction (str): 'positive' or 'negative' spread direction
        
    Returns:
        list: List of opportunity dictionaries
    """
    opportunities = []
    in_opportunity = False
    start_idx = None
    periods_below_threshold = 0

    for idx, (timestamp, is_opp) in enumerate(zip(data['datetime'], opportunity_mask)):
        if is_opp and not in_opportunity:
            # Start of new opportunity
            in_opportunity = True
            start_idx = idx
            periods_below_threshold = 0
        elif is_opp and in_opportunity:
            # Continue opportunity - reset the below-threshold counter
            periods_below_threshold = 0
        elif not is_opp and in_opportunity:
            # Below threshold while in opportunity - start counting
            periods_below_threshold += 1
            
            # Only exit if we've been below threshold for enough consecutive periods
            if periods_below_threshold > exit_delay_periods:
                # End of opportunity (exit was delayed)
                in_opportunity = False
                
                # Record the opportunity (end at the point where delay period started)
                end_idx = idx - periods_below_threshold + 1 # Go back to where we first went below threshold
                opportunity_data = data.iloc[start_idx:end_idx]
                
                if len(opportunity_data) > 0:
                    opportunities.append(_create_opportunity_record(
                        opportunity_data, ticker, exchange1_name, exchange2_name, 
                        threshold_pct, direction
                    ))
                
                periods_below_threshold = 0
                
    
    # Handle case where opportunity continues to the end
    if in_opportunity and start_idx is not None:
        # If we're still in opportunity but had some periods below threshold,
        # end at the point where the delay period would have started
        if periods_below_threshold > 0:
            end_idx = len(data) - periods_below_threshold
            opportunity_data = data.iloc[start_idx:end_idx]
        else:
            opportunity_data = data.iloc[start_idx:]
            
        if len(opportunity_data) > 0:
            opportunities.append(_create_opportunity_record(
                opportunity_data, ticker, exchange1_name, exchange2_name,
                threshold_pct, direction
            ))
    
    return opportunities

def _create_opportunity_record(opportunity_data, ticker, exchange1_name, exchange2_name, threshold_pct, direction):
    """Create a standardized opportunity record"""
    
    # Reset index to avoid iloc issues
    opp_data = opportunity_data.reset_index(drop=True)

    # Determine the side of the trade => short min rate vs long max rate
    higher_exchange = exchange2_name
    lower_exchange = exchange1_name
    direction_desc = f"Short {exchange1_name}, Long {exchange2_name}"
    if opp_data[f'fr_{exchange1_name}'].iloc[0] > opp_data[f'fr_{exchange2_name}'].iloc[0]:
        higher_exchange = exchange1_name
        lower_exchange = exchange2_name
        direction_desc = f"Short {exchange2_name}, Long {exchange1_name}"
    
    # Find max absolute normalized spread safely
    abs_norm_spread = opp_data['norm_spread'].abs()
    max_norm_idx = abs_norm_spread.idxmax()
    max_norm_value = opp_data['norm_spread'].iloc[max_norm_idx]
    
    return {
        'ticker': ticker,
        'starting_date': opp_data['datetime'].iloc[0],
        'ending_date': opp_data['datetime'].iloc[-1],
        'duration_hours': ((opp_data['datetime'].iloc[-1] - opp_data['datetime'].iloc[0]).total_seconds() / 3600) + 1,
        'spread_direction': direction,
        'trading_direction': direction_desc,
        'min_fr_t0': opp_data[f'fr_{lower_exchange}'][0],
        'max_fr_t0': opp_data[f'fr_{higher_exchange}'][0],
        'expected_pnl_pct': (opp_data[f'fr_{higher_exchange}'].sum() - opp_data[f'fr_{lower_exchange}'].sum()) * 100,
        'avg_spread_pct': opp_data['spread_pct'].mean(),
        'sum_spread_pct': opp_data['spread_pct'].sum(),
        'max_spread_pct': opp_data['spread_pct'].max(),
        'min_spread_pct': opp_data['spread_pct'].min(),
        'avg_norm_spread': opp_data['norm_spread'].mean(),
        'max_norm_spread': max_norm_value,
        f'avg_fr_{exchange1_name}_pct': opp_data[f'fr_{exchange1_name}'].mean() * 100,
        f'avg_fr_{exchange2_name}_pct': opp_data[f'fr_{exchange2_name}'].mean() * 100,
        f'sum_fr_{exchange1_name}_pct': opp_data[f'fr_{exchange1_name}'].sum() * 100,
        f'sum_fr_{exchange2_name}_pct': opp_data[f'fr_{exchange2_name}'].sum() * 100,
        'threshold_used_pct': threshold_pct,
        'data_points': len(opp_data)
    }

def compute_pair_statistics(sync_data, ticker, exchange1_name, exchange2_name):
    """
    Compute comprehensive statistics for a synchronized pair
    
    Args:
        sync_data (pd.DataFrame): Synchronized data
        ticker (str): Ticker symbol
        exchange1_name, exchange2_name (str): Exchange names
        
    Returns:
        dict: Statistics dictionary
    """
    if len(sync_data) == 0:
        return None
    
    spread = sync_data['spread']
    spread_abs = sync_data['spread_abs']
    spread_pct = sync_data['spread_pct']
    spread_rel = sync_data['spread_relative']
    
    # Basic statistics
    stats_dict = {
        'ticker': ticker,
        'pair_name': f"{ticker}_{exchange1_name}_{exchange2_name}",
        'synchronized_points': len(sync_data),
        'date_range_start': sync_data['datetime'].min(),
        'date_range_end': sync_data['datetime'].max(),
        'avg_fr_exchange1': sync_data[f'fr_{exchange1_name}'].mean(),
        'avg_fr_exchange2': sync_data[f'fr_{exchange2_name}'].mean(),
        
        # Spread statistics
        'mean_spread': spread.mean(),
        'mean_spread_abs': spread_abs.mean(),
        'std_spread': spread.std(),
        'std_spread_abs': spread_abs.std(),
        'min_spread': spread.min(),
        'max_spread': spread.max(),
        'median_spread_abs': spread_abs.median(),
        
        # Percentage spread statistics
        'mean_spread_pct': spread_pct.mean(),
        'std_spread_pct': spread_pct.std(),
        'max_spread_pct': spread_pct.max(),
        'median_spread_pct': spread_pct.median(),
        
        # Relative spread statistics
        'mean_spread_relative': spread_rel.mean() if not spread_rel.isna().all() else 0,
        'max_spread_relative': spread_rel.max() if not spread_rel.isna().all() else 0,
        
        # Distribution metrics
        'skewness_spread': stats.skew(spread),
        'kurtosis_spread': stats.kurtosis(spread)
    }
    
    # Normalized spread analysis
    if spread.std() > 0:
        normalized_spread = (spread_abs - spread_abs.mean()) / spread_abs.std()
        stats_dict.update({
            'count_above_1std': sum(normalized_spread > 1),
            'count_above_2std': sum(normalized_spread > 2),
            'pct_above_1std': sum(normalized_spread > 1) / len(normalized_spread) * 100,
            'pct_above_2std': sum(normalized_spread > 2) / len(normalized_spread) * 100,
        })
    
    # Percentage bin analysis
    bins = [0.25, 0.5, 1.0, 1.5, 2.0, 3.0, 5.0]
    for bin_threshold in bins:
        count_above = sum(spread_pct >= bin_threshold)
        pct_above = count_above / len(spread_pct) * 100 if len(spread_pct) > 0 else 0
        
        stats_dict[f'count_spread_above_{bin_threshold}pct'] = count_above
        stats_dict[f'pct_spread_above_{bin_threshold}pct'] = pct_above
    
    return stats_dict

def calculate_consecutive_durations(opportunities, timestamps):
    """Calculate duration of consecutive opportunities"""
    durations = []
    current_duration_start = None
    
    for is_opportunity, timestamp in zip(opportunities, timestamps):
        if is_opportunity and current_duration_start is None:
            current_duration_start = timestamp
        elif not is_opportunity and current_duration_start is not None:
            duration = (timestamp - current_duration_start).total_seconds() / 3600
            durations.append(duration)
            current_duration_start = None
    
    # Handle case where opportunity continues to the end
    if current_duration_start is not None:
        duration = (timestamps.iloc[-1] - current_duration_start).total_seconds() / 3600
        durations.append(duration)
        
    return durations if durations else [0]

def analyze_opportunity_duration(sync_data, ticker, threshold_pct=0.5):
    """
    Analyze duration of spread opportunities
    
    Args:
        sync_data (pd.DataFrame): Synchronized data
        ticker (str): Ticker symbol
        threshold_pct (float): Percentage threshold
        
    Returns:
        dict: Duration statistics
    """
    if len(sync_data) == 0:
        return None
    
    sync_data_sorted = sync_data.sort_values('datetime')
    spread_pct = sync_data_sorted['spread_pct']
    opportunities = spread_pct > threshold_pct
    
    duration_stats = {
        'ticker': ticker,
        'threshold_pct': threshold_pct
    }
    
    if sum(opportunities) > 0:
        durations = calculate_consecutive_durations(opportunities, sync_data_sorted['datetime'])
        duration_stats.update({
            'opportunity_count': len(durations),
            'total_opportunity_time_hours': sum(durations),
            'avg_duration_hours': np.mean(durations),
            'max_duration_hours': max(durations),
            'min_duration_hours': min(durations),
            'median_duration_hours': np.median(durations),
            'opportunity_frequency_per_day': len(durations) / ((sync_data_sorted['datetime'].max() - sync_data_sorted['datetime'].min()).days + 1)
        })
    else:
        duration_stats.update({
            'opportunity_count': 0,
            'total_opportunity_time_hours': 0,
            'avg_duration_hours': 0,
            'max_duration_hours': 0,
            'min_duration_hours': 0,
            'median_duration_hours': 0,
            'opportunity_frequency_per_day': 0
        })
    
    return duration_stats

def analyze_all_pairs(df1, df2, exchange1_name, exchange2_name, common_tickers):
    """
    Analyze all pairs and return results
    
    Returns:
        tuple: (pair_stats_df, duration_stats_df, raw_data_dict)
    """
    pair_results = []
    duration_results = []
    raw_data_dict = {}
    
    print(f"\nAnalyzing {len(common_tickers)} pairs...")
    
    for ticker in common_tickers:
        print(f"\nProcessing {ticker}...")
        
        # Synchronize data
        sync_data = synchronize_pair_data(df1, df2, ticker, exchange1_name, exchange2_name)
        if len(sync_data) > 0:
            raw_data_dict[ticker] = sync_data
            
            # Compute statistics
            pair_stats = compute_pair_statistics(sync_data, ticker, exchange1_name, exchange2_name)
            if pair_stats:
                pair_results.append(pair_stats)
            
            # Duration analysis
            duration_stats = analyze_opportunity_duration(sync_data, ticker)
            if duration_stats:
                duration_results.append(duration_stats)
    
    pair_df = pd.DataFrame(pair_results) if pair_results else pd.DataFrame()
    duration_df = pd.DataFrame(duration_results) if duration_results else pd.DataFrame()
    
    return pair_df, duration_df, raw_data_dict

def create_summary_sheet(writer, pair_df, duration_df, exchange1_name, exchange2_name):
    """Create organized summary sheet"""
    summary_data = []
    
    if not pair_df.empty:
        for _, row in pair_df.iterrows():
            ticker = row['ticker']
            
            # Get duration data
            duration_data = duration_df[duration_df['ticker'] == ticker] if not duration_df.empty else pd.DataFrame()
            
            summary_row = {
                'Pair': ticker,
                'Exchange_1': exchange1_name,
                'Exchange_2': exchange2_name,
                'Synchronized_Points': int(row['synchronized_points']),
                'Date_Start': row['date_range_start'].strftime('%Y-%m-%d') if pd.notna(row['date_range_start']) else '',
                'Date_End': row['date_range_end'].strftime('%Y-%m-%d') if pd.notna(row['date_range_end']) else '',
                
                # Funding Rate Statistics
                f'Avg_FR_{exchange1_name}_pct': round(row['avg_fr_exchange1'] * 100, 4),
                f'Avg_FR_{exchange2_name}_pct': round(row['avg_fr_exchange2'] * 100, 4),
                'Mean_Spread_pct': round(row['mean_spread_pct'], 4),
                'Std_Spread_pct': round(row['std_spread_pct'], 4),
                'Max_Spread_pct': round(row['max_spread_pct'], 4),
                'Median_Spread_pct': round(row['median_spread_pct'], 4),
                
                # Opportunity Statistics
                'Count_Above_0.25pct': int(row['count_spread_above_0.25pct']),
                'Count_Above_0.5pct': int(row['count_spread_above_0.5pct']),
                'Count_Above_1.0pct': int(row['count_spread_above_1.0pct']),
                'Count_Above_1.5pct': int(row['count_spread_above_1.5pct']),
                
                'Pct_Time_Above_0.25pct': round(row['pct_spread_above_0.25pct'], 2),
                'Pct_Time_Above_0.5pct': round(row['pct_spread_above_0.5pct'], 2),
                'Pct_Time_Above_1.0pct': round(row['pct_spread_above_1.0pct'], 2),
                'Pct_Time_Above_1.5pct': round(row['pct_spread_above_1.5pct'], 2),
            }
            
            # Add duration statistics
            if not duration_data.empty:
                duration_row = duration_data.iloc[0]
                summary_row.update({
                    'Opportunity_Count': int(duration_row['opportunity_count']),
                    'Avg_Duration_Hours': round(duration_row['avg_duration_hours'], 2),
                    'Max_Duration_Hours': round(duration_row['max_duration_hours'], 2),
                    'Total_Opportunity_Hours': round(duration_row['total_opportunity_time_hours'], 2),
                    'Frequency_Per_Day': round(duration_row['opportunity_frequency_per_day'], 2)
                })
            else:
                summary_row.update({
                    'Opportunity_Count': 0,
                    'Avg_Duration_Hours': 0,
                    'Max_Duration_Hours': 0,
                    'Total_Opportunity_Hours': 0,
                    'Frequency_Per_Day': 0
                })
            
            summary_data.append(summary_row)
    
    # Create and write summary DataFrame
    summary_df = pd.DataFrame(summary_data)
    
    if not summary_df.empty:
        summary_df = summary_df.sort_values('Mean_Spread_pct', ascending=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=2)
        
        # Format the sheet
        workbook = writer.book
        summary_sheet = writer.sheets['Summary']
        
        # Add title
        summary_sheet['A1'] = f'Cross-Exchange Funding Rate Analysis: {exchange1_name} vs {exchange2_name}'
        summary_sheet['A1'].font = Font(size=14, bold=True)
        summary_sheet['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        # Format headers
        for col in range(1, len(summary_df.columns) + 1):
            cell = summary_sheet.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width

def analyze_all_opportunities(raw_data_dict, exchange1_name, exchange2_name, thresholds=[0.1, 0.25, 0.5, 0.75, 1.0], exit_delay_periods=[0]):
    """
    Analyze arbitrage opportunities for all pairs across multiple thresholds
    
    Args:
        raw_data_dict (dict): Dictionary with ticker -> synchronized data
        exchange1_name, exchange2_name (str): Exchange names
        thresholds (list): List of percentage thresholds to analyze
        exit_delay_periods (list): List of exit delay periods to test
        
    Returns:
        dict: Dictionary with threshold -> opportunities DataFrame
    """
    print(f"\n{'='*80}")
    print("ANALYZING ARBITRAGE OPPORTUNITIES WITH MULTIPLE EXIT DELAYS")
    print(f"Exit delays to test: {exit_delay_periods}")
    print(f"Thresholds to test: {thresholds}")
    print(f"{'='*80}")
    
    all_opportunities_by_delay = {}
    
    for exit_delay in exit_delay_periods:
        print(f"\n{'='*60}")
        print(f"PROCESSING EXIT DELAY: {exit_delay} PERIODS")
        print(f"{'='*60}")
        
        delay_opportunities = {}
        
        for threshold in thresholds:
            print(f"\nAnalyzing {threshold}% threshold with {exit_delay}-period exit delay...")
            threshold_opportunities = []
            
            for ticker, sync_data in raw_data_dict.items():
                print(f"  Processing {ticker} for {threshold}% threshold...")
                
                opps = identify_arbitrage_opportunities(
                    sync_data, ticker, exchange1_name, exchange2_name, 
                    threshold, exit_delay
                )
                if len(opps) > 0:
                    # Add exit_delay info to each opportunity
                    opps['exit_delay_used'] = exit_delay
                    threshold_opportunities.append(opps)
                    print(f"    Found {len(opps)} opportunities")
                else:
                    print(f"    No opportunities found")
            
            if threshold_opportunities:
                combined_opps = pd.concat(threshold_opportunities, ignore_index=True)
                # Sort by starting_date descending (latest first)
                combined_opps = combined_opps.sort_values('starting_date', ascending=False)
                delay_opportunities[threshold] = combined_opps
                
                print(f"  Total opportunities for {threshold}%: {len(combined_opps)}")
            else:
                delay_opportunities[threshold] = pd.DataFrame()
                print(f"  No opportunities found for {threshold}% threshold")
        
        all_opportunities_by_delay[exit_delay] = delay_opportunities
        
        # Summary for this exit delay
        total_opps_this_delay = sum(len(df) for df in delay_opportunities.values() if not df.empty)
        print(f"\n📊 SUMMARY FOR {exit_delay}-PERIOD EXIT DELAY: {total_opps_this_delay} total opportunities")
    
    # Overall summary across all delays
    print(f"\n{'='*80}")
    print("OVERALL SUMMARY ACROSS ALL EXIT DELAYS")
    print(f"{'='*80}")
    
    for exit_delay in exit_delay_periods:
        delay_data = all_opportunities_by_delay[exit_delay]
        total_opps = sum(len(df) for df in delay_data.values() if not df.empty)
        
        if total_opps > 0:
            all_delay_opps = pd.concat([df for df in delay_data.values() if not df.empty], ignore_index=True)
            avg_duration = all_delay_opps['duration_hours'].mean()
            max_duration = all_delay_opps['duration_hours'].max()
            
            print(f"\nExit Delay {exit_delay} periods:")
            print(f"  Total Opportunities: {total_opps}")
            print(f"  Average Duration: {avg_duration:.2f} hours")
            print(f"  Max Duration: {max_duration:.2f} hours")
        else:
            print(f"\nExit Delay {exit_delay} periods: No opportunities found")
    
    return all_opportunities_by_delay

def create_opportunities_sheets(writer, all_opportunities_by_delay, exchange1_name, exchange2_name):
    """
    Create Excel sheets for arbitrage opportunities across different thresholds and exit delays
    
    Args:
        writer: Excel writer object
        all_opportunities_by_delay (dict): Nested dict {exit_delay: {threshold: opportunities_df}}
        exchange1_name, exchange2_name (str): Exchange names
    """
    print("Creating arbitrage opportunities sheets for multiple exit delays...")
    
    for exit_delay, delay_opportunities in all_opportunities_by_delay.items():
        print(f"\n  Processing exit delay: {exit_delay} periods")
        
        for threshold, opportunities_df in delay_opportunities.items():
            if opportunities_df.empty:
                print(f"    Skipping {threshold}% - no opportunities found")
                continue
                
            # Create sheet name with both threshold and exit delay
            sheet_name = f"Opps_{threshold}pct_Delay{exit_delay}"
            print(f"    Creating sheet: {sheet_name}")
            
            # Prepare display DataFrame with proper column names
            display_df = opportunities_df.copy()
        
            # Rename columns for clarity and match your requirements
            column_mapping = {
                'starting_date': 'Starting_Date',
                'ending_date': 'Ending_Date',
                'duration_hours': 'Duration_Hours',
                'spread_direction': 'Spread_Direction',
                'trading_direction': 'Trading_Direction',
                'min_fr_t0': 'Min_FR_t0',
                'max_fr_t0': 'Max_FR_t0',
                'expected_pnl_pct': 'Expected_Pnl_pct',
                'avg_spread_pct': 'Avg_Spread_Pct',
                'min_spread_pct': 'Min_Spread_Pct',
                'max_spread_pct': 'Max_Spread_Pct',
                'avg_norm_spread': 'Avg_Norm_Spread',
                'max_norm_spread': 'Max_Norm_Spread',
                f'avg_fr_{exchange1_name}_pct': f'Avg_FR_{exchange1_name}_Pct',
                f'avg_fr_{exchange2_name}_pct': f'Avg_FR_{exchange2_name}_Pct',
                f'sum_fr_{exchange1_name}_pct': f'Sum_FR_{exchange1_name}_Pct',
                f'sum_fr_{exchange2_name}_pct': f'Sum_FR_{exchange2_name}_Pct',
                'threshold_used_pct': 'Threshold_Pct',
                'data_points': 'Data_Points'
            }
        
            display_df = display_df.rename(columns=column_mapping)
        
            # Round numerical columns
            numeric_cols = ['Duration_Hours', 'Avg_Spread_Pct', 'Max_Spread_Pct', 'Min_Spread_Pct', 
                        'Avg_Norm_Spread', 'Max_Norm_Spread',
                        'Min_FR_t0', 'Max_FR_t0', "Expected_Pnl_pct"
                        f'Avg_FR_{exchange1_name}_Pct', f'Avg_FR_{exchange2_name}_Pct',
                            'Sum_FR_{exchange1_name}_Pct', f'Sum_FR_{exchange2_name}_Pct' ]
            for col in numeric_cols:
                if col in display_df.columns:
                    display_df[col] = display_df[col].round(6)

            # Remove timezone from all datetime columns before writing to Excel
            if display_df['Starting_Date'].dt.tz is not None: display_df['Starting_Date'] = display_df['Starting_Date'].dt.tz_localize(None)
            if display_df['Ending_Date'].dt.tz is not None: display_df['Ending_Date'] = display_df['Ending_Date'].dt.tz_localize(None)
        
            # Write to Excel
            display_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the sheet
            worksheet = writer.sheets[sheet_name]
            
            # Format headers with orange background
            for col in range(1, len(display_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
            
            # Add summary statistics at the bottom
            summary_start_row = len(display_df) + 3
            
            worksheet.cell(row=summary_start_row, column=1, value=f"SUMMARY FOR {threshold}% THRESHOLD").font = Font(size=12, bold=True)
            
            # Calculate summary statistics
            positive_opps = display_df[display_df['Spread_Direction'] == 'positive']
            negative_opps = display_df[display_df['Spread_Direction'] == 'negative']
            
            summary_stats = [
                ('Total Opportunities', len(display_df)),
                ('Positive Spread Opportunities', len(positive_opps)),
                ('Negative Spread Opportunities', len(negative_opps)),
                ('Total Duration (Hours)', display_df['Duration_Hours'].sum()),
                ('Average Duration (Hours)', display_df['Duration_Hours'].mean()),
                ('Max Duration (Hours)', display_df['Duration_Hours'].max()),
                ('Average Absolute Spread (%)', display_df['Avg_Spread_Pct'].abs().mean()),
                ('Max Absolute Spread (%)', display_df[['Max_Spread_Pct', 'Min_Spread_Pct']].abs().max().max()),
                ('Unique Tickers', display_df['ticker'].nunique()),
                ('Most Active Ticker', display_df['ticker'].value_counts().index[0] if len(display_df) > 0 else 'N/A'),
            ]
            
            for i, (stat_name, stat_value) in enumerate(summary_stats):
                worksheet.cell(row=summary_start_row + 2 + i, column=1, value=stat_name)
                if isinstance(stat_value, (int, float)):
                    worksheet.cell(row=summary_start_row + 2 + i, column=2, value=round(stat_value, 4) if isinstance(stat_value, float) else stat_value)
                else:
                    worksheet.cell(row=summary_start_row + 2 + i, column=2, value=stat_value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                worksheet.column_dimensions[column_letter].width = adjusted_width

def display_opportunities_summary_multi_delay(all_opportunities_by_delay):
    """Display summary of arbitrage opportunities across multiple exit delays"""
    print(f"\n{'='*80}")
    print("ARBITRAGE OPPORTUNITIES SUMMARY - MULTIPLE EXIT DELAYS")
    print(f"{'='*80}")
    
    if not all_opportunities_by_delay:
        print("No opportunities found.")
        return
    
    # Summary by exit delay
    for exit_delay, delay_opportunities in all_opportunities_by_delay.items():
        print(f"\n{'='*60}")
        print(f"EXIT DELAY: {exit_delay} PERIODS")
        print(f"{'='*60}")
        
        if not any(not df.empty for df in delay_opportunities.values()):
            print("No opportunities found for this exit delay.")
            continue
        
        print(f"\n📊 OPPORTUNITIES BY THRESHOLD:")
        print("-" * 50)
        
        for threshold, opps_df in delay_opportunities.items():
            if not opps_df.empty:
                total_opps = len(opps_df)
                positive_opps = len(opps_df[opps_df['spread_direction'] == 'positive'])
                negative_opps = len(opps_df[opps_df['spread_direction'] == 'negative'])
                unique_tickers = opps_df['ticker'].nunique()
                avg_duration = opps_df['duration_hours'].mean()
                avg_spread = opps_df['avg_spread_pct'].abs().mean()
                
                print(f"  {threshold}%: {total_opps} opportunities "
                      f"(+{positive_opps}, -{negative_opps}), "
                      f"Avg duration: {avg_duration:.2f}h, "
                      f"Avg spread: {avg_spread:.4f}%")
    
    # Comparison across exit delays
    print(f"\n{'='*80}")
    print("COMPARISON ACROSS EXIT DELAYS")
    print(f"{'='*80}")
    
    comparison_data = []
    for exit_delay, delay_opportunities in all_opportunities_by_delay.items():
        all_delay_opps = pd.concat([df for df in delay_opportunities.values() if not df.empty], ignore_index=True)
        
        if not all_delay_opps.empty:
            comparison_data.append({
                'exit_delay': exit_delay,
                'total_opportunities': len(all_delay_opps),
                'avg_duration': all_delay_opps['duration_hours'].mean(),
                'max_duration': all_delay_opps['duration_hours'].max(),
                'unique_tickers': all_delay_opps['ticker'].nunique()
            })
    
    if comparison_data:
        print(f"\n{'Exit Delay':<12} {'Total Opps':<12} {'Avg Duration':<15} {'Max Duration':<15} {'Unique Tickers':<15}")
        print("-" * 75)
        for data in comparison_data:
            print(f"{data['exit_delay']:<12} {data['total_opportunities']:<12} "
                  f"{data['avg_duration']:<15.2f} {data['max_duration']:<15.2f} {data['unique_tickers']:<15}")
    
    print(f"\n💡 TIP: Compare sheets 'Opps_X%_DelayY' to see impact of different exit delays!")

def create_pair_sheet(writer, ticker, sync_data, pair_df, duration_df, exchange1_name, exchange2_name):
    """Create individual pair sheet with synchronized data and statistics"""
    sheet_name = ticker[:25] if len(ticker) > 25 else ticker
    
    # Prepare synchronized data
    sync_data_display = sync_data[[
        'datetime', f'fr_{exchange1_name}', f'fr_{exchange2_name}',
        'spread', 'spread_abs', 'spread_pct'
    ]].copy()
    
    # Rename and format columns
    sync_data_display.columns = [
        'DateTime',
        f'FR_{exchange1_name}_pct',
        f'FR_{exchange2_name}_pct', 
        'Spread_Raw',
        'Spread_Abs',
        'Spread_Pct'
    ]
    
    # Convert to percentage
    sync_data_display[f'FR_{exchange1_name}_pct'] *= 100
    sync_data_display[f'FR_{exchange2_name}_pct'] *= 100
    #sync_data_display['Spread_Raw'] *= 100
    #sync_data_display['Spread_Abs'] *= 100
    
    # Round for display
    numeric_cols = [f'FR_{exchange1_name}_pct', f'FR_{exchange2_name}_pct', 
                   'Spread_Raw', 'Spread_Abs', 'Spread_Pct']
    sync_data_display[numeric_cols] = sync_data_display[numeric_cols].round(6)

    # Remove timezone from all datetime columns before writing to Excel
    if sync_data_display['DateTime'].dt.tz is not None: sync_data_display['DateTime'] = sync_data_display['DateTime'].dt.tz_localize(None)
    
    # Write to Excel
    sync_data_display.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    
    # Format headers
    for col in range(1, len(sync_data_display.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Add statistics section
    stats_start_row = len(sync_data_display) + 3
    worksheet.cell(row=stats_start_row, column=1, value=f"CROSS-EXCHANGE STATISTICS: {ticker}").font = Font(size=14, bold=True)
    
    # Get statistics
    pair_stats = pair_df[pair_df['ticker'] == ticker]
    duration_stats = duration_df[duration_df['ticker'] == ticker] if not duration_df.empty else pd.DataFrame()
    
    if not pair_stats.empty:
        stats = pair_stats.iloc[0]
        current_row = stats_start_row + 2
        
        # Funding Rate Statistics
        worksheet.cell(row=current_row, column=1, value="FUNDING RATE STATISTICS").font = Font(bold=True, color="0000FF")
        current_row += 1
        
        fr_stats = [
            ('Data Points Synchronized', int(stats['synchronized_points'])),
            ('Date Range', f"{stats['date_range_start'].strftime('%Y-%m-%d')} to {stats['date_range_end'].strftime('%Y-%m-%d')}"),
            (f'Average FR {exchange1_name} (%)', round(stats['avg_fr_exchange1'] * 100, 4)),
            (f'Average FR {exchange2_name} (%)', round(stats['avg_fr_exchange2'] * 100, 4)),
            ('Mean Spread (%)', round(stats['mean_spread_pct'], 4)),
            ('Std Deviation Spread (%)', round(stats['std_spread_pct'], 4)),
            ('Max Spread (%)', round(stats['max_spread_pct'], 4)),
            ('Median Spread (%)', round(stats['median_spread_pct'], 4)),
        ]
        
        for stat_name, stat_value in fr_stats:
            worksheet.cell(row=current_row, column=1, value=stat_name)
            worksheet.cell(row=current_row, column=2, value=stat_value)
            current_row += 1
        
        # Opportunity Statistics
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="OPPORTUNITY STATISTICS").font = Font(bold=True, color="0000FF")
        current_row += 1
        
        opp_stats = [
            ('Count Spread > 0.25%', int(stats['count_spread_above_0.25pct'])),
            ('Count Spread > 0.5%', int(stats['count_spread_above_0.5pct'])),
            ('Count Spread > 1.0%', int(stats['count_spread_above_1.0pct'])),
            ('Count Spread > 1.5%', int(stats['count_spread_above_1.5pct'])),
            ('', ''),
            ('% Time Spread > 0.25%', round(stats['pct_spread_above_0.25pct'], 2)),
            ('% Time Spread > 0.5%', round(stats['pct_spread_above_0.5pct'], 2)),
            ('% Time Spread > 1.0%', round(stats['pct_spread_above_1.0pct'], 2)),
            ('% Time Spread > 1.5%', round(stats['pct_spread_above_1.5pct'], 2)),
        ]
        
        for stat_name, stat_value in opp_stats:
            worksheet.cell(row=current_row, column=1, value=stat_name)
            worksheet.cell(row=current_row, column=2, value=stat_value)
            current_row += 1
        
        # Duration Statistics
        if not duration_stats.empty:
            duration_data = duration_stats.iloc[0]
            current_row += 1
            worksheet.cell(row=current_row, column=1, value="OPPORTUNITY DURATION STATISTICS").font = Font(bold=True, color="0000FF")
            current_row += 1
            
            duration_stats_list = [
                ('Number of Opportunities', int(duration_data['opportunity_count'])),
                ('Average Duration (hours)', round(duration_data['avg_duration_hours'], 2)),
                ('Max Duration (hours)', round(duration_data['max_duration_hours'], 2)),
                ('Total Opportunity Time (hours)', round(duration_data['total_opportunity_time_hours'], 2)),
                ('Frequency per Day', round(duration_data['opportunity_frequency_per_day'], 2)),
            ]
            
            for stat_name, stat_value in duration_stats_list:
                worksheet.cell(row=current_row, column=1, value=stat_name)
                worksheet.cell(row=current_row, column=2, value=stat_value)
                current_row += 1
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_excel_report(df1, df2, exchange1_name, exchange2_name, filename=None):
    """
    Main function to create comprehensive Excel report
    
    Args:
        df1, df2 (pd.DataFrame): Exchange dataframes
        exchange1_name, exchange2_name (str): Exchange names
        filename (str): Output filename
        
    Returns:
        tuple: (filename, pair_df, duration_df, raw_data_dict)
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'cross_exchange_funding_analysis_{exchange1_name}_{exchange2_name}_{timestamp}.xlsx'
    
    # Get common tickers and analyze all pairs
    common_tickers = get_common_tickers(df1, df2)
    
    if not common_tickers:
        print("ERROR: No common tickers found. Cannot proceed with analysis.")
        return None, pd.DataFrame(), pd.DataFrame(), {}
    
    pair_df, duration_df, raw_data_dict = analyze_all_pairs(df1, df2, exchange1_name, exchange2_name, common_tickers)

    # Step 2: Analyze arbitrage opportunities using the synchronized data
    all_opportunities = analyze_all_opportunities(raw_data_dict, exchange1_name, exchange2_name)
    
    # Create Excel report
    print("\nCreating Excel report...")
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Summary sheet
        print("Creating summary sheet...")
        create_summary_sheet(writer, pair_df, duration_df, exchange1_name, exchange2_name)

        # Arbitrage opportunities sheets (one per threshold)
        create_opportunities_sheets(writer, all_opportunities, exchange1_name, exchange2_name)
        
        # Individual pair sheets
        for ticker in common_tickers:
            if ticker in raw_data_dict:
                print(f"Creating cross-exchange sheet for {ticker}...")
                create_pair_sheet(writer, ticker, raw_data_dict[ticker], pair_df, duration_df, exchange1_name, exchange2_name)
    
    print(f"✅ Excel report saved as: {filename}")

    # Display opportunities
    display_opportunities_summary_multi_delay(all_opportunities)
    
    # Display console summary
    display_console_summary(pair_df, exchange1_name, exchange2_name)
    
    return filename, pair_df, duration_df, raw_data_dict

def display_console_summary(pair_df, exchange1_name, exchange2_name):
    """Display summary in console with historical data depth information"""
    print("\n" + "="*100)
    print(f"CROSS-EXCHANGE FUNDING RATE ANALYSIS SUMMARY")
    print(f"{exchange1_name} vs {exchange2_name}")
    print("="*100)
    
    if not pair_df.empty:
        print(f"\n📊 ANALYZED {len(pair_df)} PAIRS (INDIVIDUAL HISTORICAL DEPTH)")
        print("-" * 80)
        
        # Data depth summary
        print(f"\n📈 HISTORICAL DATA SUMMARY:")
        print(f"Total synchronized points across all pairs: {pair_df['synchronized_points'].sum():,}")
        print(f"Average points per pair: {pair_df['synchronized_points'].mean():.0f}")
        print(f"Max points for single pair: {pair_df['synchronized_points'].max():,}")
        print(f"Min points for single pair: {pair_df['synchronized_points'].min():,}")
        print(f"Pairs with >1000 points: {sum(pair_df['synchronized_points'] > 1000)}")
        print(f"Pairs with >5000 points: {sum(pair_df['synchronized_points'] > 5000)}")
        
        # Top 5 pairs by data depth
        print(f"\n📊 TOP 5 PAIRS BY DATA POINTS:")
        top_5_data = pair_df.nlargest(5, 'synchronized_points')[
            ['ticker', 'synchronized_points', 'mean_spread_pct', 'max_spread_pct']
        ].round(4)
        print(top_5_data.to_string(index=False))
        
        # Top 5 pairs by average spread
        print(f"\n🏆 TOP 5 PAIRS BY AVERAGE SPREAD:")
        top_5_spread = pair_df.nlargest(5, 'mean_spread_pct')[
            ['ticker', 'mean_spread_pct', 'max_spread_pct', 'synchronized_points', 'pct_spread_above_0.5pct']
        ].round(4)
        print(top_5_spread.to_string(index=False))
        
        # Overall statistics
        print(f"\n📈 OVERALL STATISTICS:")
        print(f"Average spread across all pairs: {pair_df['mean_spread_pct'].mean():.4f}%")
        print(f"Maximum spread observed: {pair_df['max_spread_pct'].max():.4f}%")
        print(f"Average opportunities >0.5%: {pair_df['pct_spread_above_0.5pct'].mean():.2f}%")
        print(f"Average opportunities >1.0%: {pair_df['pct_spread_above_1.0pct'].mean():.2f}%")
        
        # Date range summary  
        print(f"\n📅 DATE RANGE SUMMARY:")
        earliest_start = pair_df['date_range_start'].min()
        latest_end = pair_df['date_range_end'].max()
        print(f"Earliest data point: {earliest_start}")
        print(f"Latest data point: {latest_end}")
        print(f"Total time span covered: {(latest_end - earliest_start).days} days")

def plot_pair_analysis(sync_data, ticker, exchange1_name, exchange2_name, figsize=(15, 12)):
    """Create comprehensive plots for a specific pair"""
    if len(sync_data) == 0:
        print(f"No synchronized data available for {ticker}")
        return
        
    fig, axes = plt.subplots(3, 2, figsize=figsize)
    fig.suptitle(f'Cross-Exchange Analysis: {ticker} ({exchange1_name} vs {exchange2_name})', fontsize=16)
    
    # Funding rate comparison over time
    axes[0, 0].plot(sync_data['datetime'], sync_data[f'fr_{exchange1_name}'] * 100, 
                   label=exchange1_name, alpha=0.7)
    axes[0, 0].plot(sync_data['datetime'], sync_data[f'fr_{exchange2_name}'] * 100, 
                   label=exchange2_name, alpha=0.7)
    axes[0, 0].set_xlabel('Date')
    axes[0, 0].set_ylabel('Funding Rate (%)')
    axes[0, 0].set_title('Funding Rates Comparison')
    axes[0, 0].legend()
    axes[0, 0].grid(True, alpha=0.3)
    
    # Spread over time
    axes[0, 1].plot(sync_data['datetime'], sync_data['spread_pct'], alpha=0.7, color='red')
    axes[0, 1].axhline(y=0.5, color='orange', linestyle='--', alpha=0.7, label='0.5% threshold')
    axes[0, 1].axhline(y=1.0, color='red', linestyle='--', alpha=0.7, label='1.0% threshold')
    axes[0, 1].set_xlabel('Date')
    axes[0, 1].set_ylabel('Spread (%)')
    axes[0, 1].set_title('Spread Over Time')
    axes[0, 1].legend()
    axes[0, 1].grid(True, alpha=0.3)
    
    # Spread distribution
    axes[1, 0].hist(sync_data['spread_pct'], bins=50, alpha=0.7, edgecolor='black')
    axes[1, 0].axvline(sync_data['spread_pct'].mean(), color='red', linestyle='--', 
                      label=f'Mean: {sync_data["spread_pct"].mean():.3f}%')
    axes[1, 0].set_xlabel('Spread (%)')
    axes[1, 0].set_ylabel('Frequency')
    axes[1, 0].set_title('Spread Distribution')
    axes[1, 0].legend()
    axes[1, 0].grid(True, alpha=0.3)
    
    # Scatter plot
    axes[1, 1].scatter(sync_data[f'fr_{exchange1_name}'] * 100, 
                      sync_data[f'fr_{exchange2_name}'] * 100, alpha=0.5)
    # Add diagonal line for perfect correlation
    min_val = min(sync_data[f'fr_{exchange1_name}'].min(), sync_data[f'fr_{exchange2_name}'].min()) * 100
    max_val = max(sync_data[f'fr_{exchange1_name}'].max(), sync_data[f'fr_{exchange2_name}'].max()) * 100
    axes[1, 1].plot([min_val, max_val], [min_val, max_val], 'r--', alpha=0.7, label='Perfect correlation')
    axes[1, 1].set_xlabel(f'{exchange1_name} Funding Rate (%)')
    axes[1, 1].set_ylabel(f'{exchange2_name} Funding Rate (%)')
    axes[1, 1].set_title('Funding Rate Correlation')
    axes[1, 1].legend()
    axes[1, 1].grid(True, alpha=0.3)
    
    # Box plots for both exchanges
    box_data = [sync_data[f'fr_{exchange1_name}'] * 100, sync_data[f'fr_{exchange2_name}'] * 100]
    axes[2, 0].boxplot(box_data, labels=[exchange1_name, exchange2_name])
    axes[2, 0].set_ylabel('Funding Rate (%)')
    axes[2, 0].set_title('Funding Rate Distribution Comparison')
    axes[2, 0].grid(True, alpha=0.3)
    
    # Cumulative opportunities
    spread_sorted = np.sort(sync_data['spread_pct'])
    cumulative_pct = np.arange(1, len(spread_sorted) + 1) / len(spread_sorted) * 100
    axes[2, 1].plot(spread_sorted, 100 - cumulative_pct)
    axes[2, 1].axvline(x=0.5, color='orange', linestyle='--', alpha=0.7, label='0.5% threshold')
    axes[2, 1].axvline(x=1.0, color='red', linestyle='--', alpha=0.7, label='1.0% threshold')
    axes[2, 1].set_xlabel('Spread (%)')
    axes[2, 1].set_ylabel('Percentage of Time Above Spread')
    axes[2, 1].set_title('Cumulative Opportunity Distribution')
    axes[2, 1].legend()
    axes[2, 1].grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.show()


# Main execution functions for file-per-ticker approach
def analyze_cross_exchange_funding_from_files(exchange1_files, exchange2_files, exchange1_name="Exchange1", exchange2_name="Exchange2", output_filename=None):
    """
    Main function to analyze cross-exchange funding rate opportunities from multiple files
    
    Args:
        exchange1_files (str or list): File pattern or list of files for exchange 1
        exchange2_files (str or list): File pattern or list of files for exchange 2
        exchange1_name (str): Name for first exchange
        exchange2_name (str): Name for second exchange
        output_filename (str): Optional output filename
        
    Returns:
        tuple: (filename, pair_df, duration_df, raw_data_dict)
    """
    print("="*80)
    print("CROSS-EXCHANGE FUNDING RATE ANALYSIS")
    print("="*80)
    
    # Load data from multiple files
    df1 = load_ticker_files(exchange1_files, exchange1_name)
    df2 = load_ticker_files(exchange2_files, exchange2_name)
    
    # Create Excel report
    filename, pair_df, duration_df, raw_data_dict = create_excel_report(
        df1, df2, exchange1_name, exchange2_name, output_filename
    )
    
    return filename, pair_df, duration_df, raw_data_dict

def analyze_cross_exchange_funding_from_directories(exchange1_dir, exchange2_dir, exchange1_name="Exchange1", exchange2_name="Exchange2", output_filename=None):
    """
    Main function to analyze cross-exchange funding rate opportunities from directories
    
    Args:
        exchange1_dir (str): Directory path for exchange 1 files
        exchange2_dir (str): Directory path for exchange 2 files
        exchange1_name (str): Name for first exchange
        exchange2_name (str): Name for second exchange
        output_filename (str): Optional output filename
        
    Returns:
        tuple: (filename, pair_df, duration_df, raw_data_dict)
    """
    print("="*80)
    print("CROSS-EXCHANGE FUNDING RATE ANALYSIS")
    print("="*80)
    
    # Load data from directories
    df1 = load_exchange_files_from_directory(exchange1_dir, exchange1_name)
    df2 = load_exchange_files_from_directory(exchange2_dir, exchange2_name)
    
    # Create Excel report
    filename, pair_df, duration_df, raw_data_dict = create_excel_report(
        df1, df2, exchange1_name, exchange2_name, output_filename
    )
    
    return filename, pair_df, duration_df, raw_data_dict

def plot_top_pairs(pair_df, raw_data_dict, exchange1_name, exchange2_name, top_n=3):
    """Plot analysis for top N pairs by average spread"""
    if pair_df.empty:
        print("No data available for plotting")
        return
        
    top_pairs = pair_df.nlargest(top_n, 'mean_spread_pct')
    
    for _, row in top_pairs.iterrows():
        ticker = row['ticker']
        if ticker in raw_data_dict:
            print(f"\nCreating plots for {ticker} (Avg spread: {row['mean_spread_pct']:.4f}%)...")
            plot_pair_analysis(raw_data_dict[ticker], ticker, exchange1_name, exchange2_name)

def get_pair_summary(pair_df, ticker):
    """Get quick summary for a specific pair"""
    if pair_df.empty:
        return "No data available"
        
    pair_data = pair_df[pair_df['ticker'] == ticker]
    if pair_data.empty:
        return f"No data found for ticker: {ticker}"
    
    row = pair_data.iloc[0]
    
    summary = f"""
    PAIR SUMMARY: {ticker}
    ========================
    Synchronized Points: {int(row['synchronized_points'])}
    Date Range: {row['date_range_start'].strftime('%Y-%m-%d')} to {row['date_range_end'].strftime('%Y-%m-%d')}
    
    Average Spread: {row['mean_spread_pct']:.4f}%
    Max Spread: {row['max_spread_pct']:.4f}%
    
    Opportunities > 0.5%: {row['pct_spread_above_0.5pct']:.2f}% of time
    Opportunities > 1.0%: {row['pct_spread_above_1.0pct']:.2f}% of time
    """
    
    return summary


# Example usage functions
def main_from_file_lists():
    """Example usage with explicit file lists"""
    
    # Define file lists for each exchange
    binance_files = [
        'binance_BTCUSDT.parquet',
        'binance_ETHUSDT.parquet', 
        'binance_ADAUSDT.parquet'
    ]
    
    bybit_files = [
        'bybit_BTCUSDT.parquet',
        'bybit_ETHUSDT.parquet',
        'bybit_ADAUSDT.parquet'
    ]
    
    # Analyze cross-exchange funding rates
    filename, pair_df, duration_df, raw_data_dict = analyze_cross_exchange_funding_from_files(
        binance_files,
        bybit_files,
        'Binance',
        'Bybit'
    )
    
    # Plot top pairs
    plot_top_pairs(pair_df, raw_data_dict, 'Binance', 'Bybit', top_n=3)
    
    return filename, pair_df, duration_df, raw_data_dict

def main_from_patterns():
    """Example usage with file patterns"""
    
    # Use glob patterns to find files
    filename, pair_df, duration_df, raw_data_dict = analyze_cross_exchange_funding_from_files(
        'data/binance_*.parquet',     # All Binance files
        'data/bybit_*.parquet',       # All Bybit files
        'Binance',
        'Bybit'
    )
    
    # Plot top pairs
    plot_top_pairs(pair_df, raw_data_dict, 'Binance', 'Bybit', top_n=3)
    
    return filename, pair_df, duration_df, raw_data_dict

def main_from_directories():
    """Example usage with directories"""
    
    exchA, exchB = "hyperliquid", "lighter"
    # Load all parquet files from directories
    filename, pair_df, duration_df, raw_data_dict = analyze_cross_exchange_funding_from_directories(
        f'{exchA}_data/frs',              # Directory with Binance files
        f'{exchB}_data/frs',              # Directory with Bybit files
        exchA,
        exchB
    )
    
    # Plot top pairs
    plot_top_pairs(pair_df, raw_data_dict, exchA, exchB, top_n=5)
    
    return filename, pair_df, duration_df, raw_data_dict

def main():
    """Default main function - choose your preferred approach"""
    
    print("Choose your data loading approach:")
    print("1. File lists")
    print("2. File patterns") 
    print("3. Directories")
    
    # For demonstration, using file patterns approach
    return main_from_directories()


if __name__ == "__main__":
    # Run the analysis
    filename, pair_df, duration_df, raw_data_dict = main()